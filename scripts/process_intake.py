"""
GW Intake Form – process_intake.py

Extracts data from a GW Material Intake Form request file and populates
the correct plant tab of the Spare Parts template.

Usage:
    python process_intake.py \
        --request <request_file.xlsx> \
        --template <template_file.xlsx> \
        --plant <plant_code> \
        --output <output_file.xlsx>
"""

import argparse
import re
import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

NO_FILL = PatternFill(fill_type=None)

# Unit conversion factors
INCH_TO_CM = 2.54
LB_TO_KG   = 0.453592


# ---------------------------------------------------------------------------
# Plant configuration
# ---------------------------------------------------------------------------

PLANT_CONFIG = {
    "US30": {
        "tab": "US",
        "unit_of_length": '"',          # imperial
        # Column-index overrides (0-based) for fields with no SAP code in template
        "fixed_by_col": {
            107: "X",   # ML activity (col 108, no SAP code)
        },
        "fixed": {
            "MARA-MBRSH":         "M",
            "MARA-MTART":         "Z002",
            "MARC-TWERK":         "US30",
            "MVKE-VKORG":         "US30",
            "MVKE-VTWEG":         "00",
            "MARA-SPART":         "04",
            "MARA-MEINS":         "EA",
            "MARA-MSTAE":         "03",
            "MARM-MEINH":         "EA",
            "MARM-UMREN":         1,
            "MARM-UMREZ":         1,
            "MARM-VOLEH":         "FT3",
            "MVKE-DWERK":         "US30",
            "MLAN-ALAND":         "US",    # Tax departure Country 1
            "TATYP":              "ZTXJ",  # Tax category 1
            "MLAN-TAXM1":         "1",     # Tax classification 1
            "MVKE-KTGRM":         "20",
            "MARA-MTPOS_MARA":    "NORM",
            "MVKE-MTPOS":         "NORM",
            "MARA-TRAGR":         "0001",
            "MARC-LADGR":         "0001",
            "MARA-VABME":         1,
            "MARC-EKGRP":         "P07",
            "MARC-WEBAZ":         4,
            "MARC-MMSTA":         "03",
            "MARC-DISMM":         "PD",
            "MARC-DISPO":         "Z01",
            "MARC-DISLS":         "MB",
            "MARC-BESKZ":         "F",
            "MARC-LGPRO":         "U302",
            "MARC-LGFSB":         "U302",
            "MARC -PLIFZ":        77,       # note: SAP code has a space
            "MARC -STRGR":        "40",
            "MARC -VRMOD":        2,
            "MARC -VINT1":        180,
            "MARC-VINT2":         60,
            "MARC-MTVFP":         "ST",
            "MARA-RAUBE":         "20",
            "MBEW -BKLAS":        "2000",
            "MBEW - PEINH":       1,
            "CKMLHD-MLAST":       "2",     # Price Determination (col 109)
            "MBEW - VPRSV":       "V",
            "MBEW-XLIFO":         "X",
            "MBEW-HKMAT":         "X",
            "MARC-LOSGR":         1,
            "MBEW-EKALR":         "X",
            "MARC-AWSLS":         "000001",
            "MARA-SERIAL":        "Z001",
        },
    },
    "CA10": {
        "tab": "CA10",
        "unit_of_length": '"',
        "fixed_by_col": {
            107: "X",   # ML activity (col 108, no SAP code)
        },
        "fixed": {
            "MARA-MBRSH":         "M",
            "MARA-MTART":         "Z002",
            "MARC-TWERK":         "CA10",
            "MVKE-VKORG":         "CA10",
            "MVKE-VTWEG":         "00",
            "MARA-SPART":         "04",
            "MARA-MEINS":         "EA",
            "MARA-MSTAE":         "03",
            "MARM-MEINH":         "PCA",
            "MARM-UMREN":         1,
            "MARM-UMREZ":         1,
            "MARM-VOLEH":         "CCM",
            "MVKE-DWERK":         "CA10",
            "MLAN-ALAND":         "CA",
            "TATYP":              "CTXJ",
            "MLAN-TAXM1":         "1",
            "MVKE-KTGRM":         "20",
            "MARA-MTPOS_MARA":    "NORM",
            "MVKE-MTPOS":         "NORM",
            "MARA-TRAGR":         "0001",
            "MARC-LADGR":         "0001",
            "MARA-VABME":         1,
            "MARC-EKGRP":         "P07",
            "MARC-WEBAZ":         2,
            "MARC-MMSTA":         "03",
            "MARC-DISMM":         "PD",
            "MARC-MINBE":         1,       # Reorder Point
            "MARC-DISPO":         "Z02",
            "MARC-DISLS":         "MB",
            "MARC-BESKZ":         "F",
            "MARC-LGPRO":         "C102",
            "MARC-LGFSB":         "C102",
            # MARC -PLIFZ (Planned delivery time) = blank for CA10
            "MARC -STRGR":        "40",
            "MARC -VRMOD":        1,
            "MARC -VINT1":        999,
            "MARC-VINT2":         0,
            "MARC-MTVFP":         "ST",
            "MARA-RAUBE":         "20",
            "MARC-SERNP":         "Z001",  # Serial Number Profile - IM level
            # MARA-SERIAL (Serial Number Profile) = blank for CA10
            "MBEW -BKLAS":        "2000",
            "MBEW - PEINH":       1,
            "CKMLHD-MLAST":       "2",     # Price Determination (col 109)
            "MBEW - VPRSV":       "V",
            "MBEW-XLIFO":         "X",
            "MBEW-HKMAT":         "X",
            "MARC-LOSGR":         1,
            "MBEW-EKALR":         "X",
            "MARC-AWSLS":         "000001",
        },
    },
    "NL10": {
        "tab": "NL10",
        "unit_of_length": "CM",
        # fixed_by_col: 0-based col index overrides for fields sharing SAP codes
        # Col 57/58/59 (0-indexed 56/57/58) = Tax Country/Category/Classification 2
        # (share SAP codes with Tax 1 → FIRST-occurrence rule only writes Tax 1)
        # Col 108 (0-indexed 107) = ML activity (no SAP code in template)
        "fixed_by_col": {56: "NL", 57: "MWST", 58: 1, 107: "X"},
        "fixed": {
            "MARA-MBRSH":         "M",
            "MARA-MTART":         "Z002",
            "MARC-TWERK":         "NL10",
            "MVKE-VKORG":         "SE10",
            "MVKE-VTWEG":         "00",
            "MARA-SPART":         "04",
            "MARA-MEINS":         "EA",
            "MARA-MSTAE":         "03",
            "MARM-MEINH":         "ST",
            "MARM-UMREN":         1,
            "MARM-UMREZ":         1,
            "MARM-VOLEH":         "CCM",
            "MVKE-DWERK":         "NL10",
            "MVKE-VRKME":         "ST",    # Sales Unit
            "MLAN-ALAND":         "GB",
            "TATYP":              "MWST",
            "MLAN-TAXM1":         "1",
            "MVKE-KTGRM":         "20",
            "MARA-MTPOS_MARA":    "NORM",
            "MVKE-MTPOS":         "NORM",
            "MARA-TRAGR":         "0001",
            "MARC-LADGR":         "0001",
            "MARA-VABME":         1,
            "MARC-EKGRP":         "P07",
            "MARC-WEBAZ":         2,
            "MARC-MMSTA":         "03",
            "MARC-DISMM":         "VB",
            "MARC-MINBE":         1,       # Reorder Point
            "MARC-DISPO":         "Z02",
            "MARC-DISLS":         "MB",
            "MARC-BESKZ":         "F",
            "MARC-LGPRO":         "NL12",
            "MARC-LGFSB":         "NL12",
            # MARC -PLIFZ (Planned delivery time) = blank for NL10
            "MARC -STRGR":        "40",
            "MARC -VRMOD":        1,
            "MARC -VINT1":        999,
            "MARC-VINT2":         0,
            "MARC-MTVFP":         "ST",
            "MARA-RAUBE":         "20",
            "MARC-SERNP":         "Z001",  # Serial Number Profile - IM level
            # MARA-SERIAL (Serial Number Profile) = blank for NL10
            "MBEW -BKLAS":        "2000",
            "MBEW - PEINH":       1,
            "CKMLHD-MLAST":       "2",     # Price Determination (col 109)
            "MBEW - VPRSV":       "V",
            "MBEW-XLIFO":         "X",
            "MBEW-HKMAT":         "X",
            "MARC-LOSGR":         1,
            "MBEW-EKALR":         "X",
            "MARC-AWSLS":         "000001",
        },
    },
    "UK10": {
        "tab": "UK10",
        "unit_of_length": "CM",
        "fixed_by_col": {107: "X"},
        "fixed": {
            "MARA-MBRSH":         "M",
            "MARA-MTART":         "Z002",
            "MARC-TWERK":         "UK10",
            "MVKE-VKORG":         "UK10",
            "MVKE-VTWEG":         "00",
            "MARA-SPART":         "04",
            "MARA-MEINS":         "EA",
            "MARA-MSTAE":         "03",
            "MARM-MEINH":         "ST",
            "MARM-UMREN":         1,
            "MARM-UMREZ":         1,
            "MARM-VOLEH":         "CCM",
            "MVKE-DWERK":         "UK12",
            "MLAN-ALAND":         "GB",
            "TATYP":              "MWST",
            "MLAN-TAXM1":         "1",
            "MVKE-KTGRM":         "20",
            "MARA-MTPOS_MARA":    "NORM",
            "MVKE-MTPOS":         "NORM",
            "MARA-TRAGR":         "0001",
            "MARC-LADGR":         "0001",
            "MARA-VABME":         1,
            "MARC-EKGRP":         "P07",
            "MARC-WEBAZ":         2,
            "MARC-MMSTA":         "03",
            "MARC-DISMM":         "VB",
            "MARC-DISPO":         "Z02",
            "MARC-DISLS":         "MB",
            "MARC-BESKZ":         "F",
            "MARC-LGPRO":         "UK10",
            "MARC-LGFSB":         "UK10",
            "MARC -PLIFZ":        77,
            "MARC -STRGR":        "40",
            "MARC -VRMOD":        1,
            "MARC -VINT1":        999,
            "MARC-VINT2":         0,
            "MARC-MTVFP":         "ST",
            "MARA-RAUBE":         "20",
            "MBEW -BKLAS":        "2000",
            "MBEW - PEINH":       1,
            "CKMLHD-MLAST":       "2",     # Price Determination (col 109)
            "MBEW - VPRSV":       "V",
            "MBEW-XLIFO":         "X",
            "MBEW-HKMAT":         "X",
            "MARC-LOSGR":         1,
            "MBEW-EKALR":         "X",
            "MARC-AWSLS":         "000001",
            "MARA-SERIAL":        "Z001",
        },
    },
    "UK11": {
        "tab": "UK11",
        "unit_of_length": "CM",
        # fixed_by_col: Tax 2 (col 57/58/59 → 0-indexed 56/57/58) + ML activity (107)
        "fixed_by_col": {56: "NL", 57: "MWST", 58: 1, 107: "X"},
        "fixed": {
            "MARA-MBRSH":         "M",
            "MARA-MTART":         "Z002",
            "MARC-TWERK":         "UK11",
            "MVKE-VKORG":         "SE10",
            "MVKE-VTWEG":         "00",
            "MARA-SPART":         "04",
            "MARA-MEINS":         "EA",
            "MARA-MSTAE":         "03",
            "MARM-MEINH":         "ST",
            "MARM-UMREN":         1,
            "MARM-UMREZ":         1,
            "MARM-VOLEH":         "CCM",
            "MVKE-DWERK":         "NL10",
            "MVKE-VRKME":         "ST",    # Sales Unit
            "MLAN-ALAND":         "GB",
            "TATYP":              "MWST",
            "MLAN-TAXM1":         "1",
            "MVKE-KTGRM":         "20",
            "MARA-MTPOS_MARA":    "NORM",
            "MVKE-MTPOS":         "NORM",
            "MARA-TRAGR":         "0001",
            "MARC-LADGR":         "0001",
            "MARA-VABME":         1,
            "MARC-EKGRP":         "P07",
            "MARC-WEBAZ":         2,
            "MARC-MMSTA":         "03",
            "MARC-DISMM":         "PD",
            "MARC-DISPO":         "Z02",
            "MARC-DISLS":         "MB",
            "MARC-BESKZ":         "F",
            "MARC-LGPRO":         "UK11",
            "MARC-LGFSB":         "UK11",
            # MARC -PLIFZ (Planned delivery time) = blank for UK11
            "MARC -STRGR":        "40",
            "MARC -VRMOD":        1,
            "MARC -VINT1":        999,
            "MARC-VINT2":         0,
            "MARC-MTVFP":         "ST",
            "MARA-RAUBE":         "20",
            "MARC-SERNP":         "Z001",  # Serial Number Profile - IM level
            # MARA-SERIAL (Serial Number Profile) = blank for UK11
            "MBEW -BKLAS":        "2000",
            "MBEW - PEINH":       1,
            "CKMLHD-MLAST":       "2",     # Price Determination (col 109)
            "MBEW - VPRSV":       "V",
            "MBEW-XLIFO":         "X",
            "MBEW-HKMAT":         "X",
            "MARC-LOSGR":         1,
            "MBEW-EKALR":         "X",
            "MARC-AWSLS":         "000001",
        },
    },
    "UK12": {
        "tab": "UK12",
        "unit_of_length": "CM",
        "fixed_by_col": {107: "X"},
        "fixed": {
            "MARA-MBRSH":         "M",
            "MARA-MTART":         "Z002",
            "MARC-TWERK":         "UK12",
            "MVKE-VKORG":         "UK10",
            "MVKE-VTWEG":         "00",
            "MARA-SPART":         "04",
            "MARA-MEINS":         "EA",
            "MARA-MSTAE":         "03",
            "MARM-MEINH":         "ST",
            "MARM-UMREN":         1,
            "MARM-UMREZ":         1,
            "MARM-VOLEH":         "CCM",
            "MVKE-DWERK":         "UK12",
            "MLAN-ALAND":         "GB",
            "TATYP":              "MWST",
            "MLAN-TAXM1":         "1",
            "MVKE-KTGRM":         "20",
            "MARA-MTPOS_MARA":    "NORM",
            "MVKE-MTPOS":         "NORM",
            "MARA-TRAGR":         "0001",
            "MARC-LADGR":         "0001",
            "MARA-VABME":         1,
            "MARC-EKGRP":         "P07",
            "MARC-WEBAZ":         2,
            "MARC-MMSTA":         "03",
            "MARC-DISMM":         "VB",
            "MARC-DISPO":         "Z02",
            "MARC-DISLS":         "MB",
            "MARC-BESKZ":         "F",
            "MARC-LGPRO":         "UK12",
            "MARC-LGFSB":         "UK12",
            "MARC -PLIFZ":        77,
            "MARC -STRGR":        "40",
            "MARC -VRMOD":        1,
            "MARC -VINT1":        999,
            "MARC-VINT2":         0,
            "MARC-MTVFP":         "ST",
            "MARA-RAUBE":         "20",
            "MBEW -BKLAS":        "2000",
            "MBEW - PEINH":       1,
            "CKMLHD-MLAST":       "2",     # Price Determination (col 109)
            "MBEW - VPRSV":       "V",
            "MBEW-XLIFO":         "X",
            "MBEW-HKMAT":         "X",
            "MARC-LOSGR":         1,
            "MBEW-EKALR":         "X",
            "MARC-AWSLS":         "000001",
            "MARA-SERIAL":        "Z001",
        },
    },
}

# SAP codes for fields that need special handling
SAP_MATERIAL_CODE   = "MARA-MATNR"
SAP_UNIT_OF_LENGTH  = "MARM-MEABM"
SAP_STD_PRICE       = "MBEW - VERPR"
SAP_L10_COST        = "MBEW-ZPLP3"
SAP_L10_COST_DATE   = "MBEW-ZPLD3"

# Price table: which L8 column to use per plant
# Matched by fuzzy keyword in column header (case-insensitive, strip whitespace/\xa0)
PLANT_L8_KEYWORD = {
    "US30": "US",    # matches L8_US, L8US, L8 US
    "CA10": "CAD",   # matches L8_CAD, L8CAD, L8 CA (also catches CA)
    "NL10": "EUR",
    "UK10": "EUR",
    "UK11": "EUR",
    "UK12": "EUR",
}

# CA10 also accept "CA" if no "CAD" column found
CA10_FALLBACK_KEYWORD = "CA"

# Fields that need unit conversion (imperial → metric)
SAP_LENGTH    = "MARM-LAENG"   # × 2.54
SAP_WIDTH     = "MARM-BREIT"   # × 2.54
SAP_HEIGHT    = "MARM-HOEHE"   # × 2.54
SAP_GROSS_WT  = "MARM-BRGEW"   # × 0.453592
SAP_NET_WT    = "MARA-NTGEW"   # × 0.453592
SAP_WT_UNIT   = "MARM-GEWEI"   # → "KG"
SAP_VOLUME    = "MARM-VOLUM"   # → blank


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize_sap(code: str) -> str:
    """Normalize SAP code: strip and collapse internal spaces."""
    if not code:
        return ""
    return re.sub(r"\s+", " ", str(code).strip())


def find_sheet(wb, target_name: str):
    """Find a worksheet by name, tolerant of case and extra whitespace.

    Request files sometimes have the "Intake form" tab renamed with different
    capitalization (e.g. "Intake Form", "intake form"). An exact-match lookup
    breaks on any of these variants, so we normalize both sides before
    comparing.
    """
    target_norm = re.sub(r"\s+", " ", target_name.strip()).lower()
    for name in wb.sheetnames:
        if re.sub(r"\s+", " ", name.strip()).lower() == target_norm:
            return wb[name]
    raise KeyError(
        f"Worksheet matching '{target_name}' not found. "
        f"Available sheets: {wb.sheetnames}"
    )


def build_sap_map(ws, sap_row: int = 3) -> dict:
    """Return {normalized_sap_code: 0-based_col_index} using FIRST occurrence.

    Some SAP codes appear multiple times (e.g. MLAN-ALAND for Tax 1-4, or
    MARM-MEABM for single-unit and multi-unit sections). We always want the
    first occurrence — in the template that's Tax 1 / basic dimensions; in the
    request file that's the single-unit measurement section.
    """
    mapping = {}
    row = list(ws.iter_rows(min_row=sap_row, max_row=sap_row, values_only=True))[0]
    for idx, code in enumerate(row):
        if code:
            key = normalize_sap(str(code))
            if key not in mapping:          # keep FIRST occurrence only
                mapping[key] = idx
    return mapping


def clean_price(value) -> float | None:
    """Extract a numeric value with 2 decimal places; return None if not possible."""
    if value is None:
        return None
    s = re.sub(r"[^\d.\-]", "", str(value))
    if not s:
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def get_row_value(row: tuple, col_idx: int | None):
    """Safely get a value from a row tuple by column index."""
    if col_idx is None or col_idx >= len(row):
        return None
    return row[col_idx]


# ---------------------------------------------------------------------------
# Price table helpers
# ---------------------------------------------------------------------------

def clean_str(val) -> str:
    """Strip whitespace and non-breaking spaces from a value."""
    if val is None:
        return ""
    return re.sub(r"[\xa0\s]+", " ", str(val)).strip()


def find_col_by_keyword(headers: list, *keywords) -> int | None:
    """Return 0-based index of first header containing ALL keywords (case-insensitive)."""
    for idx, h in enumerate(headers):
        h_clean = clean_str(h).upper().replace("_", " ").replace("-", " ")
        if all(k.upper() in h_clean for k in keywords):
            return idx
    return None


def load_price_table(request_path: str, plant: str) -> dict:
    """Read Sheet1 of the request file and return {material_code: (std_price, l10_cost)}.

    Columns searched (fuzzy):
      - Material code : column containing '料件' or '料号' or 'part'
      - L10 cost      : column containing 'L10' and 'USD'
      - L8 std price  : column matching PLANT_L8_KEYWORD[plant]
    """
    wb = load_workbook(request_path, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        return {}

    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    # Find header row (first row that has 'L10' or 'L8' in any cell)
    header_idx = 0
    for i, row in enumerate(rows):
        row_str = " ".join(clean_str(v) for v in row if v is not None).upper()
        if "L10" in row_str or "L8" in row_str:
            header_idx = i
            break

    headers = [clean_str(v) for v in rows[header_idx]]

    # Locate columns
    mat_col = find_col_by_keyword(headers, "料件") or find_col_by_keyword(headers, "料号")
    if mat_col is None:
        # fallback: look for 'part' keyword
        mat_col = find_col_by_keyword(headers, "PART")
    l10_col = find_col_by_keyword(headers, "L10", "USD")

    # L8 column: use plant keyword, with fallback for CA10
    l8_keyword = PLANT_L8_KEYWORD.get(plant.upper(), "")
    l8_col = find_col_by_keyword(headers, "L8", l8_keyword) if l8_keyword else None
    if l8_col is None and plant.upper() == "CA10":
        l8_col = find_col_by_keyword(headers, "L8", CA10_FALLBACK_KEYWORD)

    if mat_col is None:
        return {}

    price_map = {}
    for row in rows[header_idx + 1:]:
        if not any(v is not None for v in row):
            continue
        mat_code = clean_str(row[mat_col]) if mat_col < len(row) else ""
        if not mat_code:
            continue
        std_price = clean_price(row[l8_col]) if l8_col is not None and l8_col < len(row) else None
        l10_cost  = clean_price(row[l10_col]) if l10_col is not None and l10_col < len(row) else None
        price_map[mat_code] = (std_price, l10_cost)

    return price_map


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def convert_value(sap: str, val, convert_units: bool):
    """Apply imperial→metric conversion for dimension/weight fields if needed."""
    if not convert_units or val is None:
        return val
    sap_n = normalize_sap(sap)
    if sap_n in (normalize_sap(SAP_LENGTH), normalize_sap(SAP_WIDTH), normalize_sap(SAP_HEIGHT)):
        try:
            return round(float(val) * INCH_TO_CM, 2)
        except (ValueError, TypeError):
            return val
    if sap_n in (normalize_sap(SAP_GROSS_WT), normalize_sap(SAP_NET_WT)):
        try:
            return round(float(val) * LB_TO_KG, 2)
        except (ValueError, TypeError):
            return val
    return val


def process(request_path: str, template_path: str, plant: str, output_path: str,
            convert_units: bool = False):
    plant = plant.upper()
    if plant not in PLANT_CONFIG:
        raise ValueError(f"Unknown plant '{plant}'. Supported: {list(PLANT_CONFIG.keys())}")

    config = PLANT_CONFIG[plant]
    tab_name = config["tab"]
    fixed_values = {normalize_sap(k): v for k, v in config["fixed"].items()}
    fixed_by_col = config.get("fixed_by_col", {})

    today_str = date.today().strftime("%Y%m%d")

    # Load price table from Sheet1 of the request file
    price_table = load_price_table(request_path, plant)
    print(f"Price table loaded: {len(price_table)} entries for plant {plant}")

    # --- Load request file ---
    req_wb = load_workbook(request_path, data_only=True)
    req_ws = find_sheet(req_wb, "Intake form")
    req_sap_map = build_sap_map(req_ws, sap_row=3)

    # Column index for unit-of-length filtering in request
    uol_col = req_sap_map.get(normalize_sap(SAP_UNIT_OF_LENGTH))

    # When converting: always take imperial rows (inch symbol)
    # When not converting: take rows matching the plant's native unit
    filter_unit = '"' if convert_units else config["unit_of_length"]

    # Collect data rows (row 10 onwards)
    all_req_rows = list(req_ws.iter_rows(min_row=10, values_only=True))

    def matches_unit(row):
        if uol_col is None:
            return True
        val = get_row_value(row, uol_col)
        if val is None:
            return True
        return str(val).strip() == filter_unit

    filtered_rows = [r for r in all_req_rows if any(c is not None for c in r) and matches_unit(r)]

    # --- Load template & copy target sheet ---
    shutil.copy2(template_path, output_path)
    Path(output_path).chmod(0o644)
    tmpl_wb = load_workbook(output_path)

    # Remove all sheets except the target tab
    for sheet in tmpl_wb.sheetnames:
        if sheet != tab_name:
            del tmpl_wb[sheet]

    tmpl_ws = tmpl_wb[tab_name]
    tmpl_sap_map = build_sap_map(tmpl_ws, sap_row=3)

    # Delete example data rows (row 10 onwards)
    max_row = tmpl_ws.max_row
    if max_row >= 10:
        tmpl_ws.delete_rows(10, max_row - 9)

    # Build reverse map: template col index -> normalized SAP code
    tmpl_col_to_sap = {v: k for k, v in tmpl_sap_map.items()}

    # Number of template columns
    n_tmpl_cols = tmpl_ws.max_column

    # Normalized SAP codes for unit-override fields (when converting)
    sap_uol_n    = normalize_sap(SAP_UNIT_OF_LENGTH)
    sap_wt_n     = normalize_sap(SAP_WT_UNIT)
    sap_vol_n    = normalize_sap(SAP_VOLUME)

    # --- Write data rows ---
    rows_written = 0
    blanks_log = []

    sap_std_price_n = normalize_sap(SAP_STD_PRICE)
    sap_l10_cost_n  = normalize_sap(SAP_L10_COST)
    sap_mat_code_n  = normalize_sap(SAP_MATERIAL_CODE)

    for req_row in filtered_rows:
        out_row = [None] * n_tmpl_cols

        # Get material code for this row (used for price table lookup)
        mc_col_req = req_sap_map.get(sap_mat_code_n)
        mat_code = clean_str(get_row_value(req_row, mc_col_req))
        prices = price_table.get(mat_code, (None, None))  # (std_price, l10_cost)

        for tmpl_col_idx in range(n_tmpl_cols):
            sap = tmpl_col_to_sap.get(tmpl_col_idx)
            if sap is None:
                continue

            # 1. Fixed value?
            if sap in fixed_values:
                out_row[tmpl_col_idx] = fixed_values[sap]
                continue

            # 2. Special: L10 Cost Date → today
            if sap == normalize_sap(SAP_L10_COST_DATE):
                out_row[tmpl_col_idx] = today_str
                continue

            # 3. Special: Old material number → same as Material Code
            if sap == normalize_sap("MARA-BISMT"):
                out_row[tmpl_col_idx] = mat_code or None
                continue

            # 4. Price table: Standard Price/MAP and L10 Cost come from Sheet1
            if sap == sap_std_price_n:
                out_row[tmpl_col_idx] = prices[0]  # L8 for this plant
                continue
            if sap == sap_l10_cost_n:
                out_row[tmpl_col_idx] = prices[1]  # L10(USD)
                continue

            # 5. Unit conversion overrides (when --convert is set)
            if convert_units:
                if sap == sap_uol_n:
                    out_row[tmpl_col_idx] = "CM"
                    continue
                if sap == sap_wt_n:
                    out_row[tmpl_col_idx] = "KG"
                    continue
                if sap == sap_vol_n:
                    out_row[tmpl_col_idx] = None   # leave volume blank
                    continue

            # 6. Extract from request by SAP code
            req_col = req_sap_map.get(sap)
            val = get_row_value(req_row, req_col)

            # 7. Apply dimension/weight conversion if needed
            val = convert_value(sap, val, convert_units)

            if val is None and req_col is None:
                blanks_log.append(sap)

            out_row[tmpl_col_idx] = val

        # Apply column-index overrides (for fields with no SAP code in template)
        for col_idx, val in fixed_by_col.items():
            if col_idx < len(out_row):
                out_row[col_idx] = val

        tmpl_ws.append(out_row)

        # Strip fill color from the newly appended row
        for cell in tmpl_ws[tmpl_ws.max_row]:
            cell.fill = NO_FILL

        rows_written += 1

    # Clear all conditional formatting (template CF rules can override cell.fill = NO_FILL)
    tmpl_ws.conditional_formatting._cf_rules.clear()

    # Final pass: strip fill from all data rows (row 10 onwards)
    for row in tmpl_ws.iter_rows(min_row=10):
        for cell in row:
            cell.fill = NO_FILL

    tmpl_wb.save(output_path)

    print(f"Done. Plant: {plant} | Tab: {tab_name} | Rows written: {rows_written}")
    if blanks_log:
        unique_blanks = sorted(set(blanks_log))
        print(f"SAP codes not found in request (left blank): {unique_blanks}")

    return rows_written


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="GW Intake Form processor")
    parser.add_argument("--request",  required=True, help="Path to request Excel file")
    parser.add_argument("--template", required=True, help="Path to template Excel file")
    parser.add_argument("--plants",   required=True, nargs="+",
                        help="One or more target plant codes (e.g. US30  or  NL10 UK11)")
    parser.add_argument("--output-dir", required=True, help="Directory to save output file(s)")
    parser.add_argument("--output-prefix", default="GW-IntakeForm",
                        help="Filename prefix (default: GW-IntakeForm)")
    parser.add_argument("--convert", action="store_true",
                        help="Convert imperial units (inch/LB) to metric (CM/KG)")
    args = parser.parse_args()

    today = date.today().strftime("%Y%m%d")
    out_dir = Path(args.output_dir)

    for plant in args.plants:
        output_path = str(out_dir / f"{args.output_prefix}-{plant.upper()}-{today}.xlsx")
        process(args.request, args.template, plant, output_path,
                convert_units=args.convert)
